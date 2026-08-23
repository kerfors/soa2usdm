"""
Index Generator

Generates collection-level index.html with navigation to all protocol
visualizations, source documents, and extraction artifacts.

Reads studies_protocols.xlsx for metadata (if available) and discovers
which protocols have source files, extraction outputs, and visualizations.
"""

import json
import html as html_lib
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from .base import PipelineStepBase
from . import config
from .corrections import review_status
from .nav import NAV_CSS, nav_block, page_title


def esc(text) -> str:
    return html_lib.escape(str(text)) if text else ""


# Rendered reports and JSON viewers regenerate on source mtime alone, so a
# nav-markup change would never reach files whose source did not change.
# The shared nav block is the marker: a file without it predates the shared
# navigation and is regenerated once.
_NAV_MARKER = '<nav class="pnav"'


def _is_current(html_path: Path, src_path: Path) -> bool:
    """True when the rendered HTML is newer than its source and already
    carries the shared navigation block."""
    return (html_path.exists()
            and html_path.stat().st_mtime >= src_path.stat().st_mtime
            and _NAV_MARKER in html_path.read_text(encoding='utf-8'))


def load_study_metadata(collection_path: Path) -> dict:
    """Load metadata from studies_protocols.xlsx if available.
    
    Returns dict: nct_id -> {d4k_folder, study_acronym, study_code, soa_pages, protocol_url, ...}
    """
    xlsx_path = collection_path / "studies_protocols.xlsx"
    if not xlsx_path.exists():
        # Check parent (for collections with /protocols/ subfolder)
        xlsx_path = collection_path.parent / "studies_protocols.xlsx"
    if not xlsx_path.exists():
        return {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        
        metadata = {}
        
        # Read 'studies' sheet
        if 'studies' in wb.sheetnames:
            ws = wb['studies']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else '' for h in rows[0]]
                for row in rows[1:]:
                    record = dict(zip(headers, row))
                    # Metadata is keyed by the protocol folder name the
                    # generator iterates over: an explicit protocol_id column
                    # wins; else nct_id (folders named by NCT ID); else
                    # d4k_folder (non-NCT studies in usdm_data, e.g. CDISC_Pilot).
                    key = (record.get('protocol_id') or record.get('nct_id')
                           or record.get('d4k_folder', ''))
                    if key:
                        metadata[key] = record
        
        # Merge 'protocols' sheet (soa_pages, protocol_url)
        if 'protocols' in wb.sheetnames:
            ws = wb['protocols']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else '' for h in rows[0]]
                for row in rows[1:]:
                    record = dict(zip(headers, row))
                    nct_id = record.get('nct_id', '')
                    if nct_id and nct_id in metadata:
                        metadata[nct_id]['soa_pages'] = record.get('soa_pages', '')
                        metadata[nct_id]['protocol_url'] = record.get('protocol_url', '')
        
        wb.close()
        return metadata
    except Exception:
        return {}


def discover_protocol_outputs(protocol_id: str, collection: str) -> dict:
    """Discover all files and outputs for a protocol."""
    collection_path = config.get_collection_path(collection)
    protocol_path = collection_path / protocol_id
    
    result = {
        'source_files': [],
        'extraction_json_files': [],
        'has_resolved': False,
        'resolved_files': [],
        'has_consolidated': False,
        'consolidated_file': None,
        'report_file': None,
        'review_file': None,
        'redacted': 0,
        'activities': 0,
    }
    
    if not protocol_path.exists():
        return result
    
    # Source files in protocol root. Only the sliced SoA PDF is published; the
    # full CSP is linked to its source URL (protocol_url) in the row renderer and
    # the markdown dump is not published, so neither is emitted as a local link.
    soa_pdf = protocol_path / f'{protocol_id}_soa.pdf'
    if soa_pdf.exists():
        result['source_files'].append({
            'filename': soa_pdf.name,
            'path': f"{protocol_id}/{soa_pdf.name}",
            'label': 'SoA PDF',
            'css_class': 'link-soa',
        })
    
    # SoA2USDM folder
    try:
        soa_folder = config.get_soa2usdm_folder(protocol_id, collection)
    except FileNotFoundError:
        return result
    
    extracted_dir = soa_folder / "extracted"
    if extracted_dir.exists():
        # Extraction JSONs in extracted/
        for ejson in sorted(extracted_dir.glob("*_extraction.json")):
            label = ejson.name.replace(f'{protocol_id}_', '').replace('_extraction.json', '').replace('Table_', 'T')
            tnum = int(label[1:])
            viewer_rel = _render_json_html(ejson, collection_path, protocol_id, collection,
                                           f"Table {tnum} extraction data",
                                           current=('extraction', tnum))
            result['extraction_json_files'].append({
                'filename': ejson.name,
                'path': viewer_rel,
                'label': label,
            })
        review = extracted_dir / f"{protocol_id}_review.html"
        if review.exists():
            result['review_file'] = f"{protocol_id}/SoA2USDM/extracted/{review.name}"
        # Open judgement calls: derived from review_items + sidecar references.
        status = review_status(extracted_dir)
        result['review_items_total'] = status['total']
        result['review_items_open'] = status['open']
        # Uncertainty report (per-protocol; .md preferred, .txt fallback).
        # Markdown is rendered to HTML because GitHub Pages serves .md as plain
        # text: the tables arrive as raw pipes and long lines do not wrap.
        for ext in ('.md', '.txt'):
            report = extracted_dir / f"{protocol_id}_uncertainty_report{ext}"
            if not report.exists():
                continue
            if ext == '.md':
                report_html = report.with_suffix('.html')
                if not _is_current(report_html, report):
                    _render_markdown_html(report, report_html, protocol_id, collection_path,
                                          collection, title="Extraction log", current="log")
                result['report_file'] = str(report_html.relative_to(collection_path))
            else:
                result['report_file'] = f"{protocol_id}/SoA2USDM/extracted/{report.name}"
            break

    # Resolved HTMLs + JSONs
    resolved_dir = soa_folder / "resolved"
    if resolved_dir.exists():
        html_files = sorted(resolved_dir.glob("*_resolved.html"))
        if html_files:
            result['has_resolved'] = True
            for hf in html_files:
                name = hf.name
                label = name.replace(f'{protocol_id}_', '').replace('_resolved.html', '').replace('Table_', 'T')
                json_name = name.replace('_resolved.html', '_resolved.json')
                json_viewer = None
                json_file = resolved_dir / json_name
                if json_file.exists():
                    json_viewer = _render_json_html(json_file, collection_path, protocol_id,
                                                    collection, f"Table {int(label[1:])} resolved data")
                result['resolved_files'].append({
                    'filename': name,
                    'path': f"{protocol_id}/SoA2USDM/resolved/{name}",
                    'label': label,
                    'json_path': json_viewer,
                })
    
    # Consolidated HTML
    cons_dir = soa_folder / "consolidated"
    if cons_dir.exists():
        cons_file = cons_dir / f"{protocol_id}_consolidated.html"
        if cons_file.exists():
            result['has_consolidated'] = True
            result['consolidated_file'] = f"{protocol_id}/SoA2USDM/consolidated/{cons_file.name}"
            
            # JSON companion viewer
            cons_json = cons_dir / f"{protocol_id}_consolidated.json"
            if cons_json.exists():
                result['consolidated_json'] = _render_json_html(cons_json, collection_path,
                                                                protocol_id, collection,
                                                                "consolidated data")
            
            # Read stats from JSON
            cons_json = cons_dir / f"{protocol_id}_consolidated.json"
            if cons_json.exists():
                try:
                    with open(cons_json) as f:
                        data = json.load(f)
                    meta = data.get('consolidation_metadata', {})
                    result['tables'] = len(meta.get('source_tables', []))
                    result['activities'] = meta.get('unified_activity_count', 0)
                    result['compression'] = meta.get('compression_percent', 0)
                    result['columns'] = sum(
                        len(cols) for cols in data.get('timeline_segments', {}).values()
                    )
                    # Redaction count: a data-quality signal for anyone
                    # selecting protocols (public documents black-bar some
                    # rows; is_redacted is exception-based in the JSON).
                    result['redacted'] = sum(
                        1 for ua in data.get('unified_activities', [])
                        if ua.get('is_redacted')
                    )
                except Exception:
                    pass
    
    # USDM readiness/evaluation markdown
    # Search protocol root, SoA2USDM folder, and consolidated/ subfolder.
    # Match: {NCTID}_USDM_readiness*.md, {NCTID}_USDM_readiness_report*.md,
    #        {NCTID}_USDM_evaluation*.md (with optional version suffix before .md)
    for eval_dir in [protocol_path, soa_folder, soa_folder / "consolidated"]:
        if not eval_dir.exists():
            continue
        eval_files = sorted([
            f for f in eval_dir.glob(f"{protocol_id}_USDM_*.md")
            if '_readiness' in f.name.lower() or '_evaluation' in f.name.lower()
        ])
        if eval_files:
            md_file = eval_files[-1]  # latest version
            html_file = md_file.with_suffix('.html')
            # Generate HTML wrapper if missing or stale
            if not _is_current(html_file, md_file):
                _render_markdown_html(md_file, html_file, protocol_id, collection_path, collection)
            rel = html_file.relative_to(collection_path)
            result['usdm_evaluation'] = str(rel)
            result['usdm_evaluation_label'] = md_file.stem  # filename without extension
            break
    
    return result


def _render_markdown_html(md_path: Path, html_path: Path, protocol_id: str, collection_path: Path,
                          collection: str, title: str = "USDM Readiness Evaluation",
                          current: str = None):
    """Render a per-protocol markdown report as styled HTML."""
    # Depth below the collection index — reports can sit at the protocol
    # root, in SoA2USDM/, or in a layer subfolder.
    rel_to_collection = html_path.relative_to(collection_path)
    depth = len(rel_to_collection.parts) - 1
    nav_html = nav_block(collection, protocol_id, title, depth=depth,
                         current=(current, None))

    try:
        import markdown
        md_text = md_path.read_text(encoding='utf-8')
        body = markdown.markdown(md_text, extensions=['tables', 'fenced_code'])
    except ImportError:
        # Fallback: basic rendering without markdown library
        md_text = md_path.read_text(encoding='utf-8')
        body = f'<pre style="white-space: pre-wrap; font-family: inherit;">{html_lib.escape(md_text)}</pre>'
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{esc(page_title(protocol_id, title, collection))}</title>
<style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
           max-width: 960px; margin: 40px auto; padding: 0 20px; color: #333; line-height: 1.6; }}
    {NAV_CSS}
    h1 {{ color: #1F4788; border-bottom: 2px solid #1F4788; padding-bottom: 8px; font-size: 22px; }}
    h2 {{ color: #2E75B6; border-bottom: 1px solid #ddd; padding-bottom: 4px; margin-top: 32px; font-size: 17px; }}
    h3 {{ color: #444; margin-top: 24px; font-size: 14px; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0; font-size: 12px; }}
    th {{ background: #f0f0f0; padding: 8px 10px; text-align: left; border: 1px solid #ddd; font-weight: 600; }}
    td {{ padding: 6px 10px; border: 1px solid #eee; vertical-align: top; }}
    tr:hover {{ background: #f8f9fa; }}
    code {{ background: #f4f4f4; padding: 2px 5px; border-radius: 3px; font-size: 12px; }}
    pre {{ background: #f8f8f8; padding: 16px; border-radius: 6px; overflow-x: auto; font-size: 12px; }}
    pre code {{ background: none; padding: 0; }}
    blockquote {{ border-left: 3px solid #6a1b9a; margin: 16px 0; padding: 8px 16px; background: #faf5fc; color: #555; }}
    strong {{ color: #222; }}
    hr {{ border: none; border-top: 1px solid #ddd; margin: 24px 0; }}
</style>
</head>
<body>
    {nav_html}
    {body}
</body>
</html>'''
    html_path.write_text(html, encoding='utf-8')


def _render_json_html(json_path: Path, collection_path: Path, protocol_id: str,
                      collection: str, page_label: str, current: tuple = (None, None)):
    """Render a JSON file as a styled, interactive HTML viewer.

    Generates {name}_viewer.html alongside the JSON file with:
    - Syntax-highlighted JSON (keys, strings, numbers, booleans, null)
    - Collapsible objects/arrays (click to expand/collapse)
    - Search/filter box
    - Shared breadcrumb + sibling strip (soa2usdm.nav)
    - Link to download raw JSON

    Only regenerates if the JSON is newer than the existing viewer HTML or
    the viewer predates the shared navigation block.
    """
    viewer_path = json_path.with_name(json_path.stem + '_viewer.html')

    # Skip if viewer is up to date
    if _is_current(viewer_path, json_path):
        return str(viewer_path.relative_to(collection_path))

    rel_to_collection = viewer_path.relative_to(collection_path)
    depth = len(rel_to_collection.parts) - 1
    nav_html = nav_block(collection, protocol_id, page_label, depth=depth, current=current)

    # Read and escape JSON content
    json_text = json_path.read_text(encoding='utf-8')
    json_escaped = html_lib.escape(json_text)

    title = f"{protocol_id} — {page_label}"

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{esc(page_title(protocol_id, page_label, collection))}</title>
<style>
    * {{ box-sizing: border-box; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
           margin: 0; padding: 20px; background: #f8f9fa; color: #333; }}
    {NAV_CSS}
    .toolbar {{ display: flex; align-items: center; gap: 12px; margin-bottom: 16px; flex-wrap: wrap; }}
    .title {{ font-size: 16px; font-weight: 600; color: #1F4788; }}
    .raw-link {{ font-size: 11px; }}
    .raw-link a {{ color: #888; text-decoration: none; font-family: monospace; }}
    .raw-link a:hover {{ color: #555; text-decoration: underline; }}
    #search {{ padding: 6px 12px; border: 1px solid #ccc; border-radius: 4px; font-size: 12px;
               width: 240px; margin-left: auto; }}
    #search:focus {{ outline: none; border-color: #2E75B6; box-shadow: 0 0 0 2px rgba(46,117,182,0.15); }}
    .stats {{ font-size: 11px; color: #888; }}
    
    #json-container {{ background: white; border: 1px solid #e0e0e0; border-radius: 8px;
                       padding: 20px; overflow: auto; max-height: calc(100vh - 120px);
                       font-family: "SF Mono", "Fira Code", "Fira Mono", Menlo, Consolas, monospace;
                       font-size: 12px; line-height: 1.6; }}
    
    /* Syntax highlighting */
    .json-key {{ color: #881391; font-weight: 500; }}
    .json-string {{ color: #1a6b2c; }}
    .json-number {{ color: #1750eb; }}
    .json-boolean {{ color: #d32f2f; font-weight: 500; }}
    .json-null {{ color: #888; font-style: italic; }}
    .json-bracket {{ color: #555; }}
    
    /* Collapsible */
    .json-toggle {{ cursor: pointer; user-select: none; position: relative; }}
    .json-toggle::before {{ content: "▼"; display: inline-block; width: 14px; font-size: 9px;
                            color: #aaa; transition: transform 0.15s; }}
    .json-toggle.collapsed::before {{ transform: rotate(-90deg); }}
    .json-toggle:hover::before {{ color: #1F4788; }}
    .json-content {{ margin-left: 20px; }}
    .json-content.hidden {{ display: none; }}
    .json-ellipsis {{ display: none; color: #aaa; font-style: italic; font-size: 11px; }}
    .json-toggle.collapsed + .json-content + .json-ellipsis {{ display: inline; }}
    .json-comma {{ color: #555; }}
    
    /* Search highlight */
    .highlight {{ background: #fff3b0; border-radius: 2px; padding: 0 1px; }}
    .search-no-match {{ opacity: 0.2; }}
</style>
</head>
<body>
    {nav_html}
    <div class="toolbar">
        <div class="title">{esc(title)}</div>
        <div class="raw-link"><a href="{json_path.name}" download>raw json</a></div>
        <div class="stats" id="stats"></div>
        <input type="text" id="search" placeholder="Search keys and values…">
    </div>
    <div id="json-container"></div>
    
    <script>
    const jsonData = {json_text};
    
    function escHtml(s) {{
        return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
    }}
    
    function renderJson(data, indent, isLast) {{
        const comma = isLast ? '' : '<span class="json-comma">,</span>';
        if (data === null) return `<span class="json-null">null</span>${{comma}}`;
        if (typeof data === 'boolean') return `<span class="json-boolean">${{data}}</span>${{comma}}`;
        if (typeof data === 'number') return `<span class="json-number">${{data}}</span>${{comma}}`;
        if (typeof data === 'string') return `<span class="json-string">"${{escHtml(data)}}"</span>${{comma}}`;
        
        const isArray = Array.isArray(data);
        const entries = isArray ? data.map((v, i) => [i, v]) : Object.entries(data);
        const open = isArray ? '[' : '{{';
        const close = isArray ? ']' : '}}';
        
        if (entries.length === 0) {{
            return `<span class="json-bracket">${{open}}${{close}}</span>${{comma}}`;
        }}
        
        const id = 'n' + Math.random().toString(36).substr(2, 8);
        let html = `<span class="json-toggle" data-target="${{id}}" onclick="toggleNode(this)">`;
        html += `<span class="json-bracket">${{open}}</span></span>`;
        html += `<div class="json-content" id="${{id}}">`;
        
        entries.forEach(([key, val], i) => {{
            const last = i === entries.length - 1;
            const keyHtml = isArray ? '' : `<span class="json-key">"${{escHtml(key)}}"</span>: `;
            html += `<div class="json-line">${{keyHtml}}${{renderJson(val, indent + 1, last)}}</div>`;
        }});
        
        html += `</div><span class="json-ellipsis"> ${{entries.length}} items… </span>`;
        html += `<span class="json-bracket">${{close}}</span>${{comma}}`;
        return html;
    }}
    
    function toggleNode(el) {{
        el.classList.toggle('collapsed');
        const target = document.getElementById(el.dataset.target);
        target.classList.toggle('hidden');
    }}
    
    // Render
    const container = document.getElementById('json-container');
    container.innerHTML = renderJson(jsonData, 0, true);
    
    // Stats
    const jsonStr = JSON.stringify(jsonData);
    const kb = (jsonStr.length / 1024).toFixed(1);
    document.getElementById('stats').textContent = `${{kb}} KB`;
    
    // Auto-collapse large arrays/objects at depth > 1
    document.querySelectorAll('.json-toggle').forEach(el => {{
        const content = document.getElementById(el.dataset.target);
        if (content && content.children.length > 20) {{
            el.classList.add('collapsed');
            content.classList.add('hidden');
        }}
    }});
    
    // Search
    const searchInput = document.getElementById('search');
    let searchTimeout;
    searchInput.addEventListener('input', function() {{
        clearTimeout(searchTimeout);
        searchTimeout = setTimeout(doSearch, 200);
    }});
    
    function doSearch() {{
        const term = searchInput.value.trim().toLowerCase();
        const lines = container.querySelectorAll('.json-line');
        
        // Clear previous highlights
        container.querySelectorAll('.highlight').forEach(el => {{
            el.outerHTML = el.textContent;
        }});
        lines.forEach(l => l.classList.remove('search-no-match'));
        
        if (!term) return;
        
        // Expand all first
        container.querySelectorAll('.json-toggle.collapsed').forEach(el => {{
            el.classList.remove('collapsed');
            document.getElementById(el.dataset.target).classList.remove('hidden');
        }});
        
        lines.forEach(line => {{
            const text = line.textContent.toLowerCase();
            if (text.includes(term)) {{
                // Highlight matches in text nodes
                const walker = document.createTreeWalker(line, NodeFilter.SHOW_TEXT);
                const textNodes = [];
                while (walker.nextNode()) textNodes.push(walker.currentNode);
                textNodes.forEach(node => {{
                    const idx = node.textContent.toLowerCase().indexOf(term);
                    if (idx >= 0) {{
                        const span = document.createElement('span');
                        span.className = 'highlight';
                        const range = document.createRange();
                        range.setStart(node, idx);
                        range.setEnd(node, idx + term.length);
                        range.surroundContents(span);
                    }}
                }});
            }} else {{
                line.classList.add('search-no-match');
            }}
        }});
    }}
    </script>
</body>
</html>'''
    viewer_path.write_text(html, encoding='utf-8')
    return str(viewer_path.relative_to(collection_path))


def generate_index_html(collection: str) -> str:
    """Generate collection index HTML."""
    collection_path = config.get_collection_path(collection)
    all_protocols = config.list_protocols(collection)
    study_meta = load_study_metadata(collection_path)
    descriptor = config.load_collection_descriptor(collection)
    # Optional provenance column, declared per collection in collection.json:
    # {"label": ..., "field": <manifest column>, "url_template": "...{value}..."}.
    # No descriptor / no provenance block -> the column is not rendered at all.
    provenance = descriptor.get('provenance')

    # Studies the manifest marks as not extractable -- e.g. the posted protocol
    # carries no SoA table. Declared as data in the manifest's excluded_reason
    # column rather than derived from the filesystem, so the list is identical
    # in a fresh clone and in a working tree that still holds the source PDFs.
    excluded = [
        {
            'nct_id': key,
            'provenance_ref': meta.get(provenance['field'], '') if provenance else '',
            'study_code': meta.get('study_code', ''),
            'study_acronym': meta.get('study_acronym', ''),
            'excluded_reason': str(meta.get('excluded_reason') or '').strip(),
        }
        for key, meta in sorted(study_meta.items())
        if str(meta.get('excluded_reason') or '').strip()
    ]
    excluded_ids = {e['nct_id'] for e in excluded}

    # Discover outputs for each protocol
    protocols = []
    for pid in all_protocols:
        if pid in excluded_ids:
            continue
        outputs = discover_protocol_outputs(pid, collection)
        meta = study_meta.get(pid, {})
        protocols.append({
            'nct_id': pid,
            'provenance_ref': meta.get(provenance['field'], '') if provenance else '',
            'study_code': meta.get('study_code', ''),
            'study_acronym': meta.get('study_acronym', ''),
            'soa_pages': meta.get('soa_pages', ''),
            'protocol_url': meta.get('protocol_url', ''),
            **outputs,
        })
    
    ready = [p for p in protocols if p['has_resolved'] or p['has_consolidated']]
    pending = [p for p in protocols if not p['has_resolved'] and not p['has_consolidated']]
    
    generated_at = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
    excl_meta = f' | {len(excluded)} excluded' if excluded else ''
    
    def protocol_row(p: dict) -> str:
        nct = p['nct_id']
        is_ready = p['has_resolved'] or p['has_consolidated']
        
        # Source column: full CSP linked to its source URL (not published here) +
        # the published sliced SoA PDF.
        source_links = []
        prot_url = p.get('protocol_url', '')
        if prot_url:
            source_links.append(
                f'<a href="{esc(prot_url)}" class="link-pdf" title="Full protocol at source" '
                f'target="_blank" rel="noopener">Protocol</a>'
            )
        for sf in p.get('source_files', []):
            source_links.append(
                f'<a href="{sf["path"]}" class="{sf["css_class"]}" title="{esc(sf["filename"])}">{sf["label"]}</a>'
            )
        source_html = ' '.join(source_links) if source_links else ''
        
        # Extraction JSON column
        ext_json_parts = []
        for ef in p.get('extraction_json_files', []):
            ext_json_parts.append(
                f'<a href="{ef["path"]}" class="link-table" title="Extraction JSON viewer — {esc(ef["filename"])}">{ef["label"]}</a>'
            )
        ext_json_html = ' '.join(ext_json_parts)
        
        # Resolved tables column
        resolved_html = ''
        if p.get('resolved_files'):
            resolved_parts = []
            for rf in p['resolved_files']:
                part = f'<a href="{rf["path"]}" class="link-table" title="Per-table view — Table {rf["label"]} with IDs and relationships">{rf["label"]}</a>'
                resolved_parts.append(part)
            resolved_html = ' '.join(resolved_parts)
        
        # Consolidated SoA column
        cons_html = ''
        if p['has_consolidated']:
            cons_html = f'<a href="{p["consolidated_file"]}" class="link-table" title="Protocol-level SoA — all tables unified">View</a>'
            if p.get('consolidated_json'):
                cons_html += f' <a href="{p["consolidated_json"]}" class="link-json" title="Consolidated JSON data">json</a>'

        # Redacted rows column: shown only where the protocol has any, as
        # "n / total" of unified activities (e.g. NCT04677179: 6 / 60).
        red_html = ''
        if p.get('redacted'):
            red_html = f'<span title="{p["redacted"]} of {p["activities"]} unified activity rows are CCI-redacted in the public protocol">{p["redacted"]} / {p["activities"]}</span>'

        # Review column: the review page, and the state of its open decisions
        # (an item is decided when a sidecar correction names it). Log column:
        # the extractor's own account of the run (the uncertainty report).
        review_html = ''
        if p.get('review_file'):
            review_html = f'<a href="{p["review_file"]}" class="link-review" title="Review page: the extraction against its source pages">review</a>'
        if p.get('review_items_total'):
            n_open, n_total = p['review_items_open'], p['review_items_total']
            state = f'{n_open} open of {n_total}' if n_open else f'all {n_total} decided'
            review_html += (f' <span class="decisions" title="Judgement calls raised by the extraction, and how many still await a reviewer">'
                            f'{state}</span>')
        report_html = ''
        if p.get('report_file'):
            report_html = f'<a href="{p["report_file"]}" class="link-report" title="The extractor\'s own account of the run: what it decided and where it was unsure">extraction log</a>'

        # USDM readiness column
        usdm_html = ''
        if p.get('usdm_evaluation'):
            usdm_label = esc(p.get('usdm_evaluation_label', 'View'))
            usdm_html = f'<a href="{p["usdm_evaluation"]}" class="link-usdm" title="USDM Readiness Evaluation">{usdm_label}</a>'
        
        # Study info
        soa = esc(str(p['soa_pages'])) if p['soa_pages'] else ''
        study_code = esc(str(p['study_code'])) if p['study_code'] else ''
        acronym = esc(str(p['study_acronym'])) if p['study_acronym'] else ''
        prov_td = ''
        if provenance:
            ref = esc(str(p['provenance_ref'])) if p['provenance_ref'] else ''
            prov_cell = (f'<a href="{esc(provenance["url_template"].replace("{value}", str(p["provenance_ref"])))}" '
                         f'target="_blank" rel="noopener">{ref}</a>'
                         if ref else '')
            prov_td = f'\n            <td class="prov-ref">{prov_cell}</td>'
        nct_cell = (f'<a href="https://clinicaltrials.gov/study/{nct}" target="_blank">{nct}</a>'
                    if nct.startswith('NCT') else nct)

        row_cls = 'ready' if is_ready else 'pending-row'

        return f'''<tr class="{row_cls}" id="{esc(nct)}">
            <td class="nct">{nct_cell}</td>{prov_td}
            <td class="study-code">{study_code}</td>
            <td class="acronym">{acronym}</td>
            <td class="soa-pages">{soa}</td>
            <td class="sources">{source_html}</td>
            <td class="viz">{ext_json_html}</td>
            <td class="viz">{resolved_html}</td>
            <td class="viz">{cons_html}</td>
            <td class="stats">{red_html}</td>
            <td class="viz">{review_html}</td>
            <td class="viz">{report_html}</td>
        </tr>'''
    
    def excluded_row(e: dict) -> str:
        prov_td = f'<td class="prov-ref">{esc(e["provenance_ref"])}</td>' if provenance else ''
        return f'''
        <tr class="excluded-row" id="{esc(e['nct_id'])}">
            <td class="nct">{esc(e['nct_id'])}</td>{prov_td}
            <td class="study-code">{esc(e['study_code'])}</td>
            <td class="acronym">{esc(e['study_acronym'])}</td>
            <td class="excluded-reason" colspan="8">Not extractable &mdash; {esc(e['excluded_reason'])}</td>
        </tr>'''

    ready_rows = ''.join(protocol_row(p) for p in ready)
    pending_rows = ''.join(protocol_row(p) for p in pending)
    excluded_rows = ''.join(excluded_row(e) for e in excluded)
    prov_th = f'<th>{esc(provenance["label"])}</th>' if provenance else ''
    
    css = """
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #f4f6f8; padding: 20px; color: #333; }
        
        .header { background: linear-gradient(135deg, #1F4788, #2E75B6); color: white; padding: 24px 30px; border-radius: 12px; margin-bottom: 20px; }
        .header .back { font-size: 11px; margin-bottom: 6px; }
        .header .back a { color: rgba(255,255,255,0.85); text-decoration: none; }
        .header .back a:hover { text-decoration: underline; }
        .header h1 { font-size: 20px; margin-bottom: 4px; }
        .header .sub { font-size: 12px; opacity: 0.8; }
        .header .meta { font-size: 11px; opacity: 0.7; margin-top: 8px; }
        .header .activities-link { display: inline-block; margin-top: 10px; padding: 6px 12px;
            border: 1px solid rgba(255,255,255,0.55); border-radius: 6px; color: #fff;
            background: rgba(255,255,255,0.08); text-decoration: none; font-size: 13px; font-weight: 600; }
        .header .activities-link:hover { background: rgba(255,255,255,0.18); }

        .section { background: white; border-radius: 8px; margin-bottom: 16px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
        .section-header { padding: 12px 20px; font-size: 13px; font-weight: 600; color: white; }
        
        .table-wrap { overflow-x: auto; }
        table { width: 100%; border-collapse: collapse; font-size: 11px; }
        th { background: #f0f0f0; padding: 8px 10px; text-align: left; font-size: 10px; font-weight: 600; color: #666; border-bottom: 2px solid #ddd; white-space: nowrap; }
        td { padding: 6px 10px; border-bottom: 1px solid #eee; vertical-align: top; }
        tr.ready:hover { background: #f8f9fa; }
        /* The row a protocol page's breadcrumb points back to (#NCT... anchor). */
        tr:target td { background: #fff3b0; }
        
        .nct { font-family: monospace; font-weight: 600; white-space: nowrap; }
        .nct a { color: #1F4788; text-decoration: none; }
        .nct a:hover { text-decoration: underline; }
        .study-code { font-family: monospace; font-size: 10px; color: #555; white-space: nowrap; }
        .prov-ref { font-family: monospace; font-size: 10px; color: #777; white-space: nowrap; }
        .acronym { font-weight: 500; color: #444; white-space: nowrap; }
        .soa-pages { white-space: nowrap; color: #888; font-family: monospace; font-size: 10px; }
        .stats { white-space: nowrap; font-family: monospace; font-size: 10px; color: #666; }
        
        .sources a, .viz a {
            display: inline-block; padding: 2px 7px; border-radius: 4px; 
            text-decoration: none; font-size: 10px; margin: 1px 2px; font-weight: 500; 
        }
        
        .link-pdf { background: #fce4ec; color: #c62828; }
        .link-pdf:hover { background: #f8bbd0; }
        .link-soa { background: #fff3e0; color: #e65100; }
        .link-soa:hover { background: #ffe0b2; }
        .link-md { background: #e8eaf6; color: #283593; }
        .link-md:hover { background: #c5cae9; }

        .link-table { background: #e8f0fe; color: #1F4788; }
        .link-table:hover { background: #d0e2fc; }
        .link-json { background: #f5f5f5; color: #888; font-size: 9px; font-family: monospace; }
        .link-json:hover { background: #e0e0e0; color: #555; }
        .link-report { background: #fff8e1; color: #8d6e00; }
        .link-report:hover { background: #ffecb3; }
        .decisions { font-size: 11px; color: #8d6e00; white-space: nowrap; }
        .link-review { background: #e8f0fa; color: #1F4788; font-weight: 600; }
        .link-review:hover { background: #d2e1f5; }
        .link-usdm { background: #f3e5f5; color: #6a1b9a; }
        .link-usdm:hover { background: #e1bee7; }
        
        .resolved-group { white-space: nowrap; color: #666; font-size: 10px; }
        .pending { color: #aaa; font-style: italic; font-size: 10px; }
        .pending-row { opacity: 0.55; }
        .excluded-row { opacity: 0.5; background: #fafafa; }
        .excluded-reason { color: #999; font-style: italic; font-size: 10px; }
    """
    
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{esc(collection)} · collection index · SoA2USDM</title>
    <style>{css}</style>
</head>
<body>
    <div class="header">
        <div class="back"><a href="../../../index.html">Collections</a> <span style="opacity:.6">›</span> {esc(collection)}</div>
        <h1>{esc(collection)}</h1>
        <div class="sub">SoA2USDM — Schedule of Activities Extraction Pipeline</div>
        <div class="meta">{len(protocols)} protocols | {len(ready)} processed | {len(pending)} pending{excl_meta} | Generated {generated_at}</div>
        <div><a href="activities.html" class="activities-link">&#9636; All extracted activities &mdash; consolidated / source-table inventory &rarr;</a></div>
    </div>
    
    <div class="section">
        <div class="section-header" style="background: #1F4788;">Protocols</div>
        <div class="table-wrap">
        <table>
            <thead><tr>
                <th>NCT ID</th>{prov_th}<th>Study Code</th><th>Acronym</th><th>SoA pp</th><th>Source</th><th title="Layer 1 — extraction JSON per table (viewer)">1. Extraction</th><th title="Layer 2 — per-table resolved: IDs, hierarchy, relationships (HTML + JSON)">2. Resolution</th><th title="Layer 3 — protocol-level unified SoA (HTML + JSON)">3. Consolidation</th><th title="Unified activity rows whose name is redacted in the public protocol (CCI) — of total unified activities">Redacted</th><th title="Review the extraction against its source pages and take the open decisions">Review</th><th title="The extractor's own account of the run (uncertainty report)">Log</th>
            </tr></thead>
            <tbody>
                {ready_rows}
                {pending_rows}
                {excluded_rows}
            </tbody>
        </table>
        </div>
    </div>
</body>
</html>'''


# =============================================================================
# Pipeline Step Class
# =============================================================================

class IndexGeneratorStep(PipelineStepBase):
    """Generate collection index.html."""
    
    step_name = "index"
    
    def execute(self, data: dict) -> dict:
        """Generate index for a collection.
        
        Args:
            data: Must contain 'source' with collection
        """
        source = data.get("source", {})
        collection = source.get("collection")
        
        if not collection:
            self._log_error("Missing collection in source")
            return {"output_file": None}
        
        try:
            html = generate_index_html(collection)
            
            collection_path = config.get_collection_path(collection)
            output_file = collection_path / "index.html"
            output_file.write_text(html, encoding="utf-8")
            
            self._analytics.increment("index_generated")
            
            return {
                "output_file": str(output_file),
                "status": "success"
            }
        
        except Exception as e:
            self._log_error(f"Index generation failed: {e}")
            return {"output_file": None, "status": "failed", "error": str(e)}
